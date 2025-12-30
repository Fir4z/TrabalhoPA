import json
import os
from datetime import datetime
import uuid
import csv
import string

import kivy

from kivy.app import App
from kivy.uix.label import Label
from kivy.uix.gridlayout import GridLayout
from kivy.uix.textinput import TextInput
from kivy.uix.button import Button
from kivy.uix.spinner import Spinner
from kivy.lang import Builder
from kivy.uix.screenmanager import ScreenManager,Screen
from kivymd.app import MDApp
from kivy.uix.popup import Popup
from kivy.uix.boxlayout import BoxLayout
from kivymd.uix.dialog import MDDialog
from kivymd.uix.button import MDRaisedButton
from kivy.clock import Clock
from kivymd.uix.dialog import MDDialog
from kivymd.uix.button import MDRaisedButton
from kivymd.uix.boxlayout import MDBoxLayout
from kivy.metrics import dp
from kivymd.uix.label import MDLabel
from kivy.config import Config
from kivy.core.window import Window
from kivy.properties import ListProperty, StringProperty
from kivy.uix.scrollview import ScrollView

from kivymd.uix.card import MDCard

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from kivy.core.window import Window

Window.maximize()


#Vamos começar a comentar isto que ta a ficar uma confusão do crl


#Variaveis necessárias
DATA_FILE = ""
CATEGORIAS = ["Pessoal", "Trabalho", "Escola"]
PRIORIDADES = ["Baixa","Média","Alta"]
ESTADOS = ["Por fazer", "Em progresso", "Concluída"]


#Gerar o id
def gerar_id():
    return str(uuid.uuid4())[:8]

#Diferentes janelas, inutil mas preciso
class FirstWindow(Screen):
    pass
#Janela inicial,login

class SecondWindow(Screen):
    pass
#Janela para adiconar tarefas

class ThirdWindow(Screen):
    pass
#Janela principal onde estão as tarefas expostas

class FourthWindow(Screen):
    pass
#Janela de edição

class FifthWindow(Screen):
    pass
#Janela ver mais, ou seja, uma tarefa exposta por completo

class WindowManager(ScreenManager):
	pass

#Só pa adicionar segurança cifra de césar
def encrypt(message, key):
    encrypted_message = ""

    letters = string.ascii_lowercase
    LETTERS = string.ascii_uppercase
    digits = string.digits

    key_letters = key % 26
    key_digits = key % 10

    for char in message:
        if char in letters:
            i = letters.index(char)
            encrypted_message += letters[(i + key_letters) % 26]

        elif char in LETTERS:
            i = LETTERS.index(char)
            encrypted_message += LETTERS[(i + key_letters) % 26]

        elif char in digits:
            i = digits.index(char)
            encrypted_message += digits[(i + key_digits) % 10]

        else:
            encrypted_message += char  # mantém espaços e símbolos

    return encrypted_message
    


#Sistema de Login
class AuthLogic:
    #O static apenas a torna numa função normal
    @staticmethod
    def logger(app):
        first = app.root.get_screen("first")
        if first.ids.user.text == "" or first.ids.password.text == "":
            app.open_popup("Aviso!","Introduza uma palavra-passe e user!")

        else:
            
            crypt = first.ids.user.text + first.ids.password.text
            user = first.ids.user.text
            crypt = encrypt(crypt,2)
            file_path = f"{crypt}.json"

            if os.path.exists(file_path):
                AuthLogic.acc(app,file_path,user)
            
            else:
                app.open_popup_confirmacao2(file_path,user)

    #Criar vários users utilizando a cifra de césar para melhor segurança
    @staticmethod
    def acc(app,file_path,user):
        global DATA_FILE
        DATA_FILE = file_path
        app.carregar_tarefas()
        app.root.current = "third"
        app.root.transition.direction = "left"
        app.atualizar_lista()
        app.open_popup("Bem vindo!",user)

        

        
            
            
            
    @staticmethod
    def clear(app):
        first = app.root.get_screen("first")
        first.ids.user.text = ""
        first.ids.password.text = ""



#Funções uteis
class FormLogic:

    #Função para alterar o tamanho da janela quando preciso (Desisti de usar n fica mt bonito)
    @staticmethod
    def size(x,y):
        Window.size = (x,y)

    #Função que coleta a informação dada para criar uma tarefa
    @staticmethod
    def submit(app):
        root = app.root.get_screen("second")
        descricao = root.ids.descricao.text
        if descricao == "":
            app.open_popup("Inválido!","Preencha pelo menos a descrição.")

        else:
            estado = root.ids.est_spinner.text
            prioridade = root.ids.prio_spinner.text
            categoria = root.ids.cat_spinner.text
            nota = root.ids.nota.text
            estado = root.ids.est_spinner.text
            app.adicionar_tarefas(descricao, estado, prioridade, categoria, nota)
            root.ids.descricao.text = ""
            root.ids.nota.text = ""
            root.ids.est_spinner.text = "Estado"
            root.ids.cat_spinner.text = "Categoria"
            root.ids.prio_spinner.text = "Prioridade"


    

#Layout das linhas das tarefas
class TarefaItem(BoxLayout):
    def __init__(self, tarefa, **kwargs):
        super().__init__(**kwargs)
        self.tarefa = tarefa
        self.orientation = "horizontal"
        self.size_hint_y = None
        self.height = 50
        #Utilizei aspas simples por causa da f-string que ja estava com aspas duplas, pode acontecer ao longo do código por vezes
        self.add_widget(Label(text=f"Descrição: {tarefa['descricao']}", size_hint_x=0.3,color=[0, 0, 0, 1]))
        self.add_widget(Label(text=f"Estado: {tarefa['estado']}", size_hint_x=0.15,color=[0, 0, 0, 1]))
        self.add_widget(Label(text=f"Prioridade: {tarefa['prioridade']}", size_hint_x=0.15,color=[0, 0, 0, 1]))
        self.add_widget(Label(text=f"Categoria: {tarefa['categoria']}", size_hint_x=0.15,color=[0, 0, 0, 1]))
        self.add_widget(Label(text=f"Data: {datetime.fromisoformat(tarefa['data_criacao']).strftime('%m/%d/%Y')}", size_hint_x=0.15, color=[0, 0, 0, 1]))
        btn_edit = Button(text="Editar", size_hint_x=0.1)
        btn_edit.bind(on_press=self.editar)
        self.add_widget(btn_edit)
        btn_remove = Button(text="Remover", size_hint_x=0.1)
        btn_remove.bind(on_press=self.remover)
        self.add_widget(btn_remove)
        btn_vermais = Button(text="Ver Mais", size_hint_x=0.1)
        btn_vermais.bind(on_press=self.vermais)
        self.add_widget(btn_vermais)

    def editar(self, instance):
        app = App.get_running_app()
        app.editar_tarefa(self.tarefa)
        app.root.current = "fourth"
        app.root.transition.direction = "left"

    def remover(self, instance):
        app = App.get_running_app()
        app.open_popup_confirmacao(self.tarefa)
    
    #Ver tarefa completa
    def vermais(self, instance):
        app = App.get_running_app()
        app.ver_tarefa(self.tarefa)
        app.root.current = "fifth"
        app.root.transition.direction = "left"




#Aplicação em si
class ToDo(MDApp):
    
    #Necessário para outra funcionar
    def __init__(self, **kwargs):
        super().__init__(**kwargs)

    tarefas = ListProperty([])
    tarefas_filtradas = ListProperty([])
    mostrando_filtradas = False

    #Carrega as tarefas do ficheiro dado .json ou cria um com o nome dado acima
    def carregar_tarefas(self):
        if os.path.exists(DATA_FILE):
            with open(DATA_FILE, "r") as f:
                self.tarefas = json.load(f)
        else:
            self.tarefas = []
        self.tarefas_filtradas = self.tarefas[:]

    #Escreve as tarefas no ficheiro
    def salvar_tarefas(self):
        with open(DATA_FILE, "w") as f:
            json.dump(self.tarefas, f, indent=4, default=str)

    #Escreve as concluidas no ficheiro excel
    def exportar_concluidas(self, nome_ficheiro="concluidas.xlsx"):
        concluidas = [t for t in self.tarefas if t.get("estado") == "Concluída"]

        if not concluidas:
            self.open_popup("Erro", "Nenhuma tarefa concluída para exportar.")
            return

        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Tarefas Concluídas"

            camposc = ["ID","Descrição","Estado","Data de Criação","Data de Conclusão","Prioridade","Categoria","Notas"]
            campos = list(concluidas[0].keys())

            # Estilos
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid")
            header_align = Alignment(horizontal="center", vertical="center")

            row_fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")

            # Cabeçalho
            for col, camposc in enumerate(camposc, 1):
                cell = ws.cell(row=1, column=col, value=camposc)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_align

            # Conteúdo
            for row, tarefa in enumerate(concluidas, 2):
                for col, campo in enumerate(campos, 1):
                    cell = ws.cell(row=row, column=col, value=tarefa.get(campo))
                    cell.fill = row_fill

            # Ajustar larguras automaticamente
            for col in range(1, len(campos) + 1):
                letra = get_column_letter(col)
                max_len = 0
                for cell in ws[letra]:
                    if cell.value:
                        max_len = max(max_len, len(str(cell.value)))
                ws.column_dimensions[letra].width = max_len + 10

            # Congelar cabeçalho
            ws.freeze_panes = "A2"

            # Ativar filtros
            ws.auto_filter.ref = ws.dimensions

            wb.save(nome_ficheiro)

            self.open_popup("Sucesso", f"Tarefas concluídas exportadas para {nome_ficheiro}")

        # Pa n dar erro
        except PermissionError:
            self.open_popup("Erro", "Não foi possível guardar o ficheiro (o ficheiro Excel está aberto?).")


    #Popup que dá pa meter oq quiser
    def open_popup(self, titulo, mensagem):
        self.dialog = MDDialog(
            title=titulo,
            text=mensagem,
            buttons=[
                MDRaisedButton(
                    text="OK",
                    on_release=lambda x: self.dialog.dismiss()
                )
            ],
            auto_dismiss=False
        )
        self.dialog.open()

    #Popup confirmação para remover
    def open_popup_confirmacao(self, tarefa):
        self.dialog = MDDialog(
            title="Remover",
            text="Tem a certeza que quer remover esta tarefa?",
            buttons=[
                MDRaisedButton(
                    text="Sim",
                    on_release=lambda x: [self.remocao(tarefa),self.dialog.dismiss()]
                ),
                MDRaisedButton(
                    text="Não",
                    on_release=lambda x: self.dialog.dismiss()
                )
        ],
            auto_dismiss=False
        )
        self.dialog.open()

    #Popup confirmação para criar novo user
    def open_popup_confirmacao2(self,file_path,user):
        self.dialog = MDDialog(
            title="Confirmação",
            text="Deseja criar um novo user ou enganou-se na palavra-passe?",
            buttons=[
                MDRaisedButton(
                    text="Criar novo user",
                    on_release=lambda x: [self.dialog.dismiss(),AuthLogic.acc(self,file_path,user)]
                ),
                MDRaisedButton(
                    text="Voltar",
                    on_release=lambda x: [AuthLogic.clear(self),self.dialog.dismiss()]
                )
        ],
            auto_dismiss=False
        )
        self.dialog.open()


    #Popup confirmação para filtrar
    def open_popup_filtrar(self):

        # Layout do conteúdo
        content = BoxLayout(
            orientation="vertical",
            spacing=15,
            size_hint_y=None,
        )
        content.bind(minimum_height=content.setter("height"))

        # Spinners
        self.spinner_estado = Spinner(
            text="Estado",
            values=ESTADOS + ["None"],
            size_hint_y=None,
            height=44,
        )

        self.spinner_categoria = Spinner(
            text="Categoria",
            values=CATEGORIAS + ["None"],
            size_hint_y=None,
            height=44,
        )

        self.spinner_prioridade = Spinner(
            text="Prioridade",
            values=PRIORIDADES + ["None"],
            size_hint_y=None,
            height=44,
        )

        # Adicionar ao layout
        content.add_widget(self.spinner_estado)
        content.add_widget(self.spinner_categoria)
        content.add_widget(self.spinner_prioridade)

        # Dialog
        self.dialog = MDDialog(
            title="Filtrar",
            type="custom",
            content_cls=content,
            buttons=[
                MDRaisedButton(
                    text="Voltar",
                    on_release=lambda x: self.dialog.dismiss()
                ),
                MDRaisedButton(
                    text="Filtrar",
                    on_release=lambda x: (self.filtrar_tarefas(),self.dialog.dismiss())
                    
                ),
            ],
            auto_dismiss=False
        )

        self.dialog.open()

    def filtrar_tarefas(self, *args):
        estado = self.spinner_estado.text
        categoria = self.spinner_categoria.text
        prioridade = self.spinner_prioridade.text

        # Valores que NÃO devem filtrar
        ignorar = {
            "estado": ("Estado", "None", None),
            "categoria": ("Categoria", "None", None),
            "prioridade": ("Prioridade", "None", None),
        }

        self.tarefas_filtradas = []

        for tarefa in self.tarefas:
            if estado not in ignorar["estado"] and tarefa.get("estado") != estado:
                continue

            if categoria not in ignorar["categoria"] and tarefa.get("categoria") != categoria:
                continue

            if prioridade not in ignorar["prioridade"] and tarefa.get("prioridade") != prioridade:
                continue

            self.tarefas_filtradas.append(tarefa)

        # Se nenhum filtro foi aplicado, mostrar tudo
        # all apenas returna true se forem todas verdadeiras
        if not self.tarefas_filtradas and all(
            v in ignorar[k]
            # o zip cria uma espécie de dicionário dos elementos dados e as posições
            for v, k in zip(
                (estado, categoria, prioridade),
                ("estado", "categoria", "prioridade")
            )
        ):
            self.mostrando_filtradas = False
            self.atualizar_lista()
        else:
            self.mostrando_filtradas = True
            self.atualizar_lista_filtrada()






        

    #Criar o dicionário da tarefa, podia ser uma class
    def adicionar_tarefas(self, descricao, estado, prioridade, categoria, nota):
        tarefa = {
            "id": gerar_id(),
            "descricao": descricao,
            "estado": estado,
            "data_criacao": datetime.now().isoformat(),
            "data_conclusao": None,
            "prioridade": prioridade,
            "categoria": categoria,
            "notas": nota
        }
        self.tarefas.append(tarefa)
        self.salvar_tarefas()
        print(self.tarefas)
        self.atualizar_lista()
        self.open_popup("Sucesso!", "Tarefa adicionada!")

    #Remover uma tarefa
    def remocao(self, tarefa):
        self.tarefas.remove(tarefa)
        self.salvar_tarefas()
        self.atualizar_lista()

    #Mostrar as cenas certas na aba edição
    def editar_tarefa(self,tarefa):
        fourth = self.root.get_screen("fourth")
        fourth.ids.descricao.text = tarefa["descricao"]
        fourth.ids.est_spinner.text = tarefa["estado"]
        fourth.ids.cat_spinner.text = tarefa["categoria"]
        fourth.ids.prio_spinner.text = tarefa["prioridade"]
        fourth.ids.nota.text = tarefa["notas"]

        #pa mandar a variavel tarefa para o kivy
        fourth.tarefa_atual = tarefa

    #salvar edição e reescrever no ficheiro 
    def salvar_edicao(self,tarefa):
        fourth = self.root.get_screen("fourth")

        if not fourth.ids.descricao.text.strip():
            self.open_popup("Erro", "Descrição não pode estar vazia.")
            return
        
        tarefa["descricao"] = fourth.ids.descricao.text 
        tarefa["estado"] = fourth.ids.est_spinner.text  
        tarefa["categoria"] = fourth.ids.cat_spinner.text  
        tarefa["prioridade"] = fourth.ids.prio_spinner.text 
        tarefa["notas"] = fourth.ids.nota.text 

        if tarefa["estado"] == "Concluída" and tarefa["data_conclusao"] is None:
            tarefa["data_conclusao"] = datetime.now().isoformat()

        self.salvar_tarefas()
        self.atualizar_lista()
        self.open_popup("Sucesso", "Tarefa editada!")
        
    #Aba para ver todos os promenores, faz com q seja visualmente mais limpo a inicial
    def ver_tarefa(self,tarefa):
        fifth = self.root.get_screen("fifth")
        fifth.ids.ID.text = ("ID: "+ tarefa["id"])
        fifth.ids.descricao.text = tarefa["descricao"]
        fifth.ids.estado.text = tarefa["estado"]
        fifth.ids.categoria.text = tarefa["categoria"]
        fifth.ids.prioridade.text = tarefa["prioridade"]
        fifth.ids.nota.text = tarefa["notas"]
        fifth.ids.criacao.text = tarefa["data_criacao"]
        #Só pq o kivy é chato...
        if tarefa["data_conclusao"] == None:
            fifth.ids.conclusao.text = "-"
        else:
            
            fifth.ids.conclusao.text = tarefa["data_conclusao"]


    #Listar
    def atualizar_lista(self):
    
        third = self.root.get_screen("third")
        task_list = third.ids.get("task_list")
        task_list.clear_widgets()
        print(f"Adding {len(self.tarefas)} tasks to task_list")
        for tarefa in self.tarefas:
            item = TarefaItem(tarefa=tarefa)
            task_list.add_widget(item)
            print(f"Added task: {tarefa['id']}")
        

    #Listar igual mas pa filtragem
    def atualizar_lista_filtrada(self):
        
        third = self.root.get_screen("third")  
        task_list = third.ids.get("task_list")  
        task_list.clear_widgets()
        print(f"Adding {len(self.tarefas)} tasks to task_list")
        for tarefa in self.tarefas_filtradas:
            item = TarefaItem(tarefa=tarefa)
            task_list.add_widget(item)
            print(f"Added task: {tarefa['id']}")
        

    #Ordenar tarefas/reescrever ficheiro
    def ordenar_tarefas(self, criterio):
        if criterio == "Prioridade":
            ordem = {"Baixa": 1, "Média": 2, "Alta": 3}
            self.tarefas.sort(key=lambda t: ordem.get(t["prioridade"], 0),reverse=True)
        elif criterio == "Data Criação":
            self.tarefas.sort(key=lambda t: t["data_criacao"])
        elif criterio == "Categoria":
            self.tarefas.sort(key=lambda t: t["categoria"])
        self.salvar_tarefas()
        self.atualizar_lista()
        self.open_popup("Sucesso", "Tarefas ordenadas!") 

    #Escolher oq se vai ordenar
    def ordenar(self):
        third = self.root.get_screen("third")
        categoria = third.ids.cate_spinner.text
        self.ordenar_tarefas(categoria)
        

    #Construtor da app
    def build(self):
        
        
        return Builder.load_file("New Window.kv")
   

#Fim
if __name__ == "__main__":
    ToDo().run()


